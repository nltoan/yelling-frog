"""
Export Module
Export crawl data to CSV, Excel, and JSON formats
"""
import csv
import json
from typing import List, Dict, Any, Optional
from pathlib import Path
import io

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from .database import Database
from .models import CrawledURL


class DataExporter:
    """Export crawl data to various formats"""

    def __init__(self, database: Database):
        """
        Args:
            database: Database instance
        """
        self.db = database

    def get_export_columns(self) -> List[str]:
        """Get list of all columns available for export"""
        return [
            # Core URL Data
            'url', 'url_encoded', 'content_type', 'status_code', 'status_text',
            'indexability', 'indexability_status', 'hash',

            # Title Data
            'title_1', 'title_1_length', 'title_1_pixel_width',
            'title_2', 'title_2_length',

            # Meta Description
            'meta_description_1', 'meta_description_1_length', 'meta_description_1_pixel_width',
            'meta_description_2',

            # Meta Keywords
            'meta_keywords_1', 'meta_keywords_1_length',

            # Headings (H1-H6)
            'h1_1', 'h1_len_1', 'h1_2', 'h1_len_2',
            'h2_1', 'h2_len_1', 'h2_2', 'h2_len_2',
            'h3_1', 'h3_len_1', 'h3_2', 'h3_len_2',
            'h4_1', 'h4_len_1', 'h4_2', 'h4_len_2',
            'h5_1', 'h5_len_1', 'h5_2', 'h5_len_2',
            'h6_1', 'h6_len_1', 'h6_2', 'h6_len_2',

            # Directives
            'meta_robots_1', 'meta_robots_2',
            'x_robots_tag_1', 'x_robots_tag_2',
            'meta_refresh_1',

            # Canonical & Pagination
            'canonical_link_element_1', 'canonical_link_element_2',
            'rel_next_1', 'rel_prev_1',
            'http_rel_next_1', 'http_rel_prev_1',

            # Performance
            'size', 'transferred', 'response_time', 'ttfb', 'last_modified',

            # Content Analysis
            'word_count', 'text_ratio', 'readability',
            'closest_similarity_match', 'closest_similarity_score',
            'no_near_duplicates', 'language',

            # Link Metrics
            'crawl_depth', 'folder_depth', 'link_score',
            'inlinks', 'unique_inlinks', 'percentage_of_total',
            'outlinks', 'unique_outlinks',
            'external_outlinks', 'unique_external_outlinks',

            # Redirect Data
            'redirect_uri', 'redirect_type', 'http_version',

            # Security Headers
            'hsts', 'hsts_value', 'csp', 'csp_value',
            'x_content_type_options', 'x_frame_options', 'x_frame_options_value',
            'referrer_policy', 'referrer_policy_value',

            # Open Graph
            'og_title', 'og_description', 'og_image', 'og_type', 'og_url',

            # Twitter Cards
            'twitter_card', 'twitter_title', 'twitter_description', 'twitter_image',

            # Structured Data
            'has_json_ld', 'has_microdata', 'has_rdfa', 'schema_types',
            'schema_validation_errors', 'schema_validation_warnings',

            # Security Issues
            'is_https', 'has_mixed_content', 'has_insecure_forms',
            'unsafe_cross_origin_links',

            # URL Issues
            'url_length', 'has_parameters', 'has_non_ascii',
            'has_underscores', 'has_uppercase',

            # Technical
            'charset', 'viewport',

            # Timestamps
            'crawled_at', 'discovered_at',

            # Issues
            'issues',
        ]

    def export_to_csv(
        self,
        session_id: str,
        output_path: str,
        columns: Optional[List[str]] = None,
        filter_name: Optional[str] = None
    ) -> int:
        """
        Export data to CSV format

        Args:
            session_id: Crawl session ID
            output_path: Path to output CSV file
            columns: List of columns to export (None = all columns)
            filter_name: Optional filter to apply (e.g., 'missing_title')

        Returns:
            Number of rows exported
        """
        # Get data
        if filter_name:
            urls = self.db.get_urls_by_filter(session_id, filter_name)
        else:
            urls = self.db.get_all_urls(session_id)

        if not urls:
            return 0

        # Default to all columns
        if not columns:
            columns = self.get_export_columns()

        # Write CSV
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns)
            writer.writeheader()

            for url_data in urls:
                row = {}
                for col in columns:
                    value = getattr(url_data, col, None)

                    # Convert lists to comma-separated strings
                    if isinstance(value, list):
                        value = ', '.join(str(v) for v in value)

                    row[col] = value

                writer.writerow(row)

        return len(urls)

    def export_to_json(
        self,
        session_id: str,
        output_path: str,
        filter_name: Optional[str] = None,
        pretty: bool = True
    ) -> int:
        """
        Export data to JSON format

        Args:
            session_id: Crawl session ID
            output_path: Path to output JSON file
            filter_name: Optional filter to apply
            pretty: Pretty-print JSON (default True)

        Returns:
            Number of records exported
        """
        # Get data
        if filter_name:
            urls = self.db.get_urls_by_filter(session_id, filter_name)
        else:
            urls = self.db.get_all_urls(session_id)

        if not urls:
            return 0

        # Convert to dict
        data = {
            'session_id': session_id,
            'total_urls': len(urls),
            'urls': [url_data.to_dict() for url_data in urls]
        }

        # Write JSON
        with open(output_path, 'w', encoding='utf-8') as jsonfile:
            if pretty:
                json.dump(data, jsonfile, indent=2, default=str)
            else:
                json.dump(data, jsonfile, default=str)

        return len(urls)

    def export_to_excel(
        self,
        session_id: str,
        output_path: str,
        columns: Optional[List[str]] = None,
        include_multiple_sheets: bool = True
    ) -> int:
        """
        Export data to Excel (.xlsx) format

        Args:
            session_id: Crawl session ID
            output_path: Path to output Excel file
            columns: List of columns to export (None = all columns)
            include_multiple_sheets: Create separate sheets for issues (default True)

        Returns:
            Number of rows exported
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        # Get data
        all_urls = self.db.get_all_urls(session_id)

        if not all_urls:
            return 0

        # Default to all columns
        if not columns:
            columns = self.get_export_columns()

        # Create workbook
        wb = openpyxl.Workbook()

        # Main sheet: All URLs
        ws_all = wb.active
        ws_all.title = "All URLs"

        # Write headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws_all.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_idx, url_data in enumerate(all_urls, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = getattr(url_data, col_name, None)

                # Convert lists to comma-separated strings
                if isinstance(value, list):
                    value = ', '.join(str(v) for v in value)

                ws_all.cell(row=row_idx, column=col_idx, value=value)

        # Create additional sheets for common filters
        if include_multiple_sheets:
            issue_filters = {
                'Missing Titles': 'missing_title',
                'Missing H1': 'missing_h1',
                'Missing Meta Desc': 'missing_meta_description',
                'Low Content': 'low_content',
                'Client Errors 4xx': 'client_error_4xx',
                'Server Errors 5xx': 'server_error_5xx',
                'Redirects': 'redirection_3xx',
                'Non-Indexable': 'non_indexable',
            }

            for sheet_name, filter_name in issue_filters.items():
                filtered_urls = self.db.get_urls_by_filter(session_id, filter_name)

                if filtered_urls:
                    ws = wb.create_sheet(title=sheet_name)

                    # Write headers
                    for col_idx, col_name in enumerate(columns, 1):
                        cell = ws.cell(row=1, column=col_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = header_font

                    # Write data
                    for row_idx, url_data in enumerate(filtered_urls, 2):
                        for col_idx, col_name in enumerate(columns, 1):
                            value = getattr(url_data, col_name, None)

                            if isinstance(value, list):
                                value = ', '.join(str(v) for v in value)

                            ws.cell(row=row_idx, column=col_idx, value=value)

        # Save workbook
        wb.save(output_path)

        return len(all_urls)

    def export_images_to_csv(
        self,
        session_id: str,
        output_path: str
    ) -> int:
        """Export images data to CSV"""
        images = self.db.get_images(session_id)

        if not images:
            return 0

        columns = [
            'page_url', 'image_url', 'alt_text', 'alt_text_length',
            'width', 'height', 'file_size',
            'missing_alt', 'missing_alt_attribute', 'missing_size_attributes'
        ]

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns)
            writer.writeheader()

            for image in images:
                row = {col: getattr(image, col, None) for col in columns}
                writer.writerow(row)

        return len(images)

    def export_links_to_csv(
        self,
        session_id: str,
        output_path: str,
        internal_only: bool = False,
        external_only: bool = False
    ) -> int:
        """Export links data to CSV"""
        links = self.db.get_links(session_id)

        if not links:
            return 0

        # Filter if needed
        if internal_only:
            links = [link for link in links if link.is_internal]
        elif external_only:
            links = [link for link in links if not link.is_internal]

        columns = ['source_url', 'target_url', 'anchor_text', 'is_internal', 'is_nofollow', 'link_type']

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns)
            writer.writeheader()

            for link in links:
                row = {col: getattr(link, col, None) for col in columns}
                writer.writerow(row)

        return len(links)

    def export_issues_report(
        self,
        session_id: str,
        output_path: str
    ) -> int:
        """
        Export comprehensive issues report to CSV

        Includes all URLs with issues categorized
        """
        all_urls = self.db.get_all_urls(session_id)

        if not all_urls:
            return 0

        # Filter to only URLs with issues
        urls_with_issues = [url for url in all_urls if url.issues]

        columns = [
            'url', 'status_code', 'indexability',
            'issues_count', 'issues_list',
            'title_1', 'meta_description_1',
            'h1_1', 'word_count'
        ]

        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=columns)
            writer.writeheader()

            for url_data in urls_with_issues:
                row = {
                    'url': url_data.url,
                    'status_code': url_data.status_code,
                    'indexability': url_data.indexability,
                    'issues_count': len(url_data.issues),
                    'issues_list': ', '.join(url_data.issues),
                    'title_1': url_data.title_1,
                    'meta_description_1': url_data.meta_description_1,
                    'h1_1': url_data.h1_1,
                    'word_count': url_data.word_count,
                }
                writer.writerow(row)

        return len(urls_with_issues)

    def get_export_stats(self, session_id: str) -> Dict[str, int]:
        """Get statistics about what can be exported"""
        stats = self.db.get_stats(session_id)

        return {
            'total_urls': stats['total_urls'],
            'images': len(self.db.get_images(session_id)),
            'links': len(self.db.get_links(session_id)),
            'issues': sum(stats.get('issues_by_type', {}).values()),
        }
